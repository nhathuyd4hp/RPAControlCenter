import logging
import os
import shutil
import subprocess
import sys
import threading
import time
import tkinter as tk
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path, PurePosixPath
from tkinter import CENTER, Button, Label

import pandas as pd
from colorama import Fore
from colorama import init as colorama_init
from office365.runtime.auth.user_credential import UserCredential
from office365.sharepoint.client_context import ClientContext
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from tkcalendar import DateEntry
from tqdm import tqdm

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")


class Kistarkoushin:
    version = "2.5"

    def __init__(self, from_date, to_date):
        self.from_date = from_date
        self.to_date = to_date

        options = Options()
        options.add_experimental_option(
            "prefs",
            {
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "safebrowsing.enabled": True,
                "profile.default_content_setting_values.automatic_downloads": 1,
                "download.default_directory": os.path.join(os.getcwd(), "Ankens"),
            },
        )
        self.driver = webdriver.Chrome(options=options)
        self.driver.maximize_window()
        self.actions = ActionChains(self.driver)

        os.makedirs("logs", exist_ok=True)
        log_filename = os.path.join("logs", datetime.now().strftime("KISTAR_bot_log_%Y%m%d_%H%M%S.log"))
        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
            handlers=[logging.FileHandler(log_filename, encoding="utf-8"), logging.StreamHandler()],
        )

    def format_date(self, date):
        date2 = datetime.strptime(str(date), "%Y-%m-%d")
        formatted_date = date2.strftime("%mÊúà%dÊó•")
        # Remove leading zero from the month
        if formatted_date[0] == "0":
            formatted_date = formatted_date[1:]
        return formatted_date

    def process_data(self):
        # Configuration
        Accessurl = "https://webaccess.nsk-cad.com/"
        folder_path = r"Ankens"
        Over = "Kistarkoushin.xlsx"
        Docs1 = "Êñ∞Ë≥áÊñô"

        def clear_excel_data(file_path: str) -> None:
            """
            Clear all data and cell styles (starting from row 2) in an Excel workbook.

            Args:
                file_path (str): Path to the Excel file to be cleared.
            """
            try:
                wb = load_workbook(file_path)
                for ws in wb.worksheets:
                    for row in ws.iter_rows(min_row=2):
                        for cell in row:
                            cell.value = None
                            cell.fill = PatternFill(fill_type=None)
                            cell.border = None
                            cell.alignment = None
                            cell.number_format = "General"

                wb.save(file_path)
                logging.info(f"üßπ Cleared all data from Excel file: {file_path}")
            except Exception as e:
                logging.error(f"‚ùå Failed to clear Excel file {file_path}: {e}")

        def Accesslogin(Accessurl):
            try:
                self.driver.get(Accessurl)
                self.driver.switch_to.window(self.driver.window_handles[0])
                time.sleep(2)

                logid = self.driver.find_element("name", "u")
                logpassword = self.driver.find_element("name", "p")
                time.sleep(1)

                logid.clear()
                logpassword.clear()
                time.sleep(1)

                logid.send_keys("NasiwakRobot")
                time.sleep(2)
                logpassword.send_keys("159753")
                time.sleep(2)

                logid.submit()
                logging.info("Successfully logged in to Webaccess")
                time.sleep(1)
            except (NoSuchElementException, TimeoutException) as e:
                logging.error(f"Failed during Accesslogin: {e}")
                return False

        def Access(from_date, to_date):
            try:
                # Navigate to the target page
                ÂèóÊ≥®‰∏ÄË¶ß_xpath = "/html/body/div[2]/div[1]/div/div/div/ul/li[4]/a"
                ÂèóÊ≥®‰∏ÄË¶ß = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, ÂèóÊ≥®‰∏ÄË¶ß_xpath)))
                ÂèóÊ≥®‰∏ÄË¶ß.click()
                time.sleep(2)

                # Reset search filters
                Reset_xpath = "/html/body/div[2]/div[2]/div[2]/form/div/div/button[2]"
                Reset = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, Reset_xpath)))
                Reset.click()
                logging.info("Reset clicked")
                time.sleep(1)

                fromD = str(from_date)
                fromDF = datetime.strptime(fromD, "%Y-%m-%d")
                f_fromD = fromDF.strftime("%Y/%m/%d")
                logging.info(f_fromD)

                toD = str(to_date)
                toDF = datetime.strptime(toD, "%Y-%m-%d")
                f_toDFD = toDF.strftime("%Y/%m/%d")
                logging.info(f_toDFD)

                # Clear the date delivery fields and enter the date range
                fromDateField = self.driver.find_element(By.NAME, "search_fix_deliver_date_from")
                fromDateField.clear()
                fromDateField.send_keys(f_fromD)
                logging.info("From date sent")
                time.sleep(1)

                toDateField = self.driver.find_element(By.NAME, "search_fix_deliver_date_to")
                toDateField.send_keys(f_toDFD)
                logging.info("To date sent")
                time.sleep(1)

                # Select the Âá∫Ëç∑Âå∫ÂàÜ dropdown options
                dropdownbtn1 = WebDriverWait(self.driver, 20).until(
                    EC.element_to_be_clickable((By.ID, "search_deliver_type_ms"))
                )
                dropdownbtn1.click()
                logging.info("Clicked on Âá∫Ëç∑Âå∫ÂàÜ")
                time.sleep(1)

                # Select Êñ∞Ë¶è
                WebDriverWait(self.driver, 10).until(
                    EC.visibility_of_all_elements_located(
                        (By.XPATH, "(//div[@class='ui-multiselect-menu ui-widget ui-widget-content ui-corner-all'])[4]")
                    )
                )
                Checkbox1 = self.driver.find_element(
                    By.XPATH,
                    "(//div[@class='ui-multiselect-menu ui-widget ui-widget-content ui-corner-all'])[4]/ul/li[2]",
                )
                Checkbox1.click()
                logging.info("Selected Êñ∞Ë¶è")
                time.sleep(1)

                # Select the Âõ≥Èù¢ dropdown options
                dropdownbtn = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.ID, "search_drawing_type_ms"))
                )
                dropdownbtn.click()
                logging.info("Clicked on Âõ≥Èù¢")
                time.sleep(1)

                # Select Ë≥áÊñôÊú™ÁùÄ
                WebDriverWait(self.driver, 10).until(
                    EC.visibility_of_all_elements_located(
                        (By.XPATH, "(//div[@class='ui-multiselect-menu ui-widget ui-widget-content ui-corner-all'])[6]")
                    )
                )
                Checkbox1 = self.driver.find_element(
                    By.XPATH,
                    "(//div[@class='ui-multiselect-menu ui-widget ui-widget-content ui-corner-all'])[6]/ul/li[2]",
                )
                Checkbox1.click()
                logging.info("Selected Ë≥áÊñôÊú™ÁùÄ")
                time.sleep(1)

                # Perform the search
                Search_xpath = "/html/body/div[2]/div[2]/div[2]/form/div/div/button[1]"
                Search = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, Search_xpath)))
                Search.click()
                time.sleep(2)

                # Set download path
                self.driver.execute_cdp_cmd(
                    "Page.setDownloadBehavior", {"behavior": "allow", "downloadPath": rf"{os.getcwd()}"}
                )

                # Download the CSV
                Download_xpath = "/html/body/div[2]/div[2]/div[2]/div[1]/a[1]"
                Download = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, Download_xpath)))
                Download.click()
                logging.info("Downloaded CSV")
                time.sleep(2)
            except (NoSuchElementException, TimeoutException) as e:
                logging.error(f"Failed during Access: {e}")
                return False

        def Kizukuprocess(Ê°à‰ª∂Áï™Âè∑, Ê°à‰ª∂Âêç, excelline, „Éì„É´„ÉÄ„ÉºÂêç):
            try:
                logging.info("Starting Kizukuprocess...")
                first_xpath = '//*[@id="menu-navbar"]/ul[1]/li[2]/a'
                WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, first_xpath))).click()
                logging.info("Clicked on first")
                time.sleep(2)

                Search_xpath = "/html/body/div/div[1]/div/section/div[1]/div/div[2]/div/div/button"
                WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, Search_xpath))).click()
                logging.info("Clicked on Search")
                time.sleep(2)

                Site_xpath = "/html/body/div[1]/div[1]/div/section/div[3]/div/div/form/div/div[2]/div/input"
                Site = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, Site_xpath)))
                Site.clear()
                Site.send_keys(Ê°à‰ª∂Âêç)
                logging.info("Input Ê°à‰ª∂Âêç")
                time.sleep(2)

                Search2_xpath = "/html/body/div[1]/div[1]/div/section/div[3]/div/div/div[2]/button[3]"
                WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, Search2_xpath))).click()
                logging.info("Clicked search")
                time.sleep(2)
                try:
                    Book_xpath = "/html/body/div/div[1]/div/section/div[2]/table/tbody/tr/td[2]/div/button[3]"
                    WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, Book_xpath))).click()
                    logging.info("Clicked Book")
                    time.sleep(2)

                    try:
                        ÊúÄÊñ∞Âõ≥Èù¢_xpath = "//*[contains(text(),'ÊúÄÊñ∞Âõ≥Èù¢')]"
                        WebDriverWait(self.driver, 10).until(
                            EC.element_to_be_clickable((By.XPATH, ÊúÄÊñ∞Âõ≥Èù¢_xpath))
                        ).click()
                        logging.info("Clicked on ÊúÄÊñ∞Âõ≥Èù¢")
                        time.sleep(2)

                        Checkall_xpath = "AllSelect"
                        WebDriverWait(self.driver, 10).until(
                            EC.element_to_be_clickable((By.NAME, Checkall_xpath))
                        ).click()
                        logging.info("Clicked on Checkall")
                        time.sleep(2)

                        # Setting the correct download path
                        download1_path = rf"{os.getcwd()}\{folder_path}\{Ê°à‰ª∂Âêç}\{Docs1}"
                        os.makedirs(download1_path, exist_ok=True)  # Ensure directory exists
                        self.driver.execute_cdp_cmd(
                            "Page.setDownloadBehavior", {"behavior": "allow", "downloadPath": download1_path}
                        )
                        logging.info(f"Set download path to: {download1_path}")

                        download_xpath = '//*[@id="download-file-submit"]/div[1]/div[1]/div/span/button[1]'
                        WebDriverWait(self.driver, 10).until(
                            EC.element_to_be_clickable((By.XPATH, download_xpath))
                        ).click()
                        logging.info("Clicked on Download")
                        time.sleep(10)

                        zip_extract(abc_path)

                        time.sleep(2)
                        logging.info(" ÊúÄÊñ∞Âõ≥Èù¢ page 1 extracted")

                        try:
                            secondpage = "/html/body/div[2]/div/div[1]/ul/li[4]/a"
                            WebDriverWait(self.driver, 10).until(
                                EC.element_to_be_clickable((By.XPATH, secondpage))
                            ).click()
                            logging.info("Clicked on 2ndpage")
                            time.sleep(2)

                            WebDriverWait(self.driver, 10).until(
                                EC.element_to_be_clickable((By.NAME, Checkall_xpath))
                            ).click()
                            logging.info("Clicked on Checkall")
                            time.sleep(2)

                            # Setting the correct download path
                            os.makedirs(download1_path, exist_ok=True)  # Ensure directory exists
                            self.driver.execute_cdp_cmd(
                                "Page.setDownloadBehavior", {"behavior": "allow", "downloadPath": download1_path}
                            )
                            logging.info(f"Set download path to: {download1_path}")

                            WebDriverWait(self.driver, 10).until(
                                EC.element_to_be_clickable((By.XPATH, download_xpath))
                            ).click()
                            logging.info("Clicked on Download")
                            time.sleep(10)
                        except Exception:
                            logging.info("No 2ndpage Found")
                        time.sleep(2)
                        self.driver.back()
                        time.sleep(2)

                        WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, Book_xpath))).click()
                        logging.info("Clicked Book")
                        time.sleep(2)
                        try:
                            „Éó„É¨„Ç´„ÉÉ„ÉàÂõ≥_xpath = "//*[contains(text(),'„Éó„É¨„Ç´„ÉÉ„ÉàÂõ≥Èù¢')]"
                            WebDriverWait(self.driver, 10).until(
                                EC.element_to_be_clickable((By.XPATH, „Éó„É¨„Ç´„ÉÉ„ÉàÂõ≥_xpath))
                            ).click()
                            logging.info("Clicked on „Éó„É¨„Ç´„ÉÉ„ÉàÂõ≥Èù¢")
                            time.sleep(2)

                            WebDriverWait(self.driver, 10).until(
                                EC.element_to_be_clickable((By.NAME, Checkall_xpath))
                            ).click()
                            logging.info("Clicked on Checkall")
                            time.sleep(2)

                            # Setting the correct download path
                            os.makedirs(download1_path, exist_ok=True)  # Ensure directory exists
                            self.driver.execute_cdp_cmd(
                                "Page.setDownloadBehavior", {"behavior": "allow", "downloadPath": download1_path}
                            )
                            logging.info(f"Set download path to: {download1_path}")

                            WebDriverWait(self.driver, 10).until(
                                EC.element_to_be_clickable((By.XPATH, download_xpath))
                            ).click()
                            logging.info("Clicked on Download")
                            time.sleep(10)

                            zip_extract(abc_path)
                            time.sleep(2)
                            logging.info(" „Éó„É¨„Ç´„ÉÉ„ÉàÂõ≥Èù¢ extracted")
                        except Exception:
                            logging.info("No „Éó„É¨„Ç´„ÉÉ„ÉàÂõ≥Èù¢ Folder Found")
                    except Exception:
                        logging.info("No ÊúÄÊñ∞Âõ≥Èù¢ Folder Found")
                        self.driver.close()
                        self.driver.switch_to.window(self.driver.window_handles[0])
                        time.sleep(1)
                        sheet[f"C{excelline}"].value = "NG"
                        return False
                    self.driver.close()
                    self.driver.switch_to.window(self.driver.window_handles[0])
                    time.sleep(2)

                    upload_path = os.path.join(os.getcwd(), folder_path, Ê°à‰ª∂Âêç, "Êñ∞Ë≥áÊñô")
                    logging.info(f"Upload path: {upload_path}")

                    fileUpload(upload_path, Ê°à‰ª∂Áï™Âè∑, Ê°à‰ª∂Âêç, „Éì„É´„ÉÄ„ÉºÂêç)
                    logging.info("Folder has been uploaded")
                    sheet[f"C{excelline}"].value = "OK"

                    Access2(Ê°à‰ª∂Áï™Âè∑)
                    logging.info("Access2")
                except Exception as e:
                    logging.info(f"Failed in Kizuku process for {Ê°à‰ª∂Âêç}: {e}")
                    sheet[f"C{excelline}"].value = "NG"
                    self.driver.close()
                    time.sleep(2)
                    return False
            except Exception as e:
                logging.error(f"Failed in Kizuku process: {e}")
                sheet[f"C{excelline}"].value = "NG"
                self.driver.close()
                return False

        def ÈÅ†ÈâÑ„Éõ„Éº„É†Login():

            # self.driver.execute_script(f"window.open('{Kizuku_url}', '_blank');")
            # self.driver.switch_to.window(self.driver.window_handles[-1])
            time.sleep(1)
            logging.info("ÈÅ†ÈâÑ„Éõ„Éº„É†")

            kstarLoginID = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, "Email")))
            kstarLoginPwd = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, "Password")))
            time.sleep(2)

            kstarLoginID.clear()
            kstarLoginPwd.clear()
            time.sleep(2)

            kstarLoginID.send_keys("haga@nsk-cad.com")
            kstarLoginPwd.send_keys("kantou1220")
            time.sleep(2)
            logging.info("Login details entered for ÈÅ†ÈâÑ„Éõ„Éº„É†")

            submit = WebDriverWait(self.driver, 30).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="authentication"]/input[4]'))
            )
            submit.click()
            time.sleep(3)
            logging.info("Login successful - ÈÅ†ÈâÑ„Éõ„Éº„É†")

        def keiaiPlanningLogin():

            time.sleep(1)
            logging.info("Keiai Planning")

            kstarLoginID = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, "Email")))
            kstarLoginPwd = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, "Password")))
            time.sleep(2)

            kstarLoginID.clear()
            kstarLoginPwd.clear()
            time.sleep(2)

            kstarLoginID.send_keys("kip@nsk-cad.com")
            kstarLoginPwd.send_keys("343nqmun")
            time.sleep(2)
            logging.info("Login details entered for KeiaiPlanning")

            submit = WebDriverWait(self.driver, 30).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="authentication"]/input[4]'))
            )
            submit.click()
            time.sleep(3)
            logging.info("Login successful - KeiaiPlanning")

        def keiaistarFudousanLogin():
            # self.driver.execute_script(f"window.open('{Kizuku_url}', '_blank');")
            # self.driver.switch_to.window(self.driver.window_handles[-1])
            time.sleep(1)
            logging.info("Keiaistar Fudousan")

            kstarLoginID = WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((By.ID, "Email")))
            kstarLoginPwd = WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((By.ID, "Password")))
            time.sleep(3)

            kstarLoginID.clear()
            kstarLoginPwd.clear()
            time.sleep(2)

            kstarLoginID.send_keys("keiai@nsk-cad.com")
            kstarLoginPwd.send_keys("nskkantou")
            time.sleep(2)
            logging.info("Login details entered for Fudosan")

            submit = WebDriverWait(self.driver, 30).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="authentication"]/input[4]'))
            )
            submit.click()
            time.sleep(3)
            logging.info("Login successful - Fudosan")

        def empty_ankens_folder():
            folder_path = "Ankens"
            if os.path.exists(folder_path):
                shutil.rmtree(folder_path)
                logging.info(f"Deleted contents of {folder_path} folder.")
            os.makedirs(folder_path)
            logging.info(f"Created an empty {folder_path} folder.")
            time.sleep(1)

        def handle_login():
            url = "https://nskkogyo.sharepoint.com/sites/2021"
            self.driver.get(url)

            username = "kushalnasiwak@nskkogyo.onmicrosoft.com"
            password = "Vay32135"

            try:
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="i0116"]'))
                ).clear()
                # Find and fill the username field
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="i0116"]'))
                ).send_keys(username)
                time.sleep(2)
                self.driver.find_element(By.XPATH, '//*[@id="idSIButton9"]').click()
                time.sleep(2)

                # Find and fill the password field
                if self.driver.find_element(By.XPATH, '//*[@id="i0118"]').text:
                    self.driver.find_element(By.XPATH, '//*[@id="i0118"]').clear()
                    time.sleep(2)
                self.driver.find_element(By.XPATH, '//*[@id="i0118"]').send_keys(password)
                time.sleep(2)
                self.driver.find_element(By.XPATH, '//*[@id="idSIButton9"]').click()
                time.sleep(2)
                self.driver.find_element(By.XPATH, '//*[@id="KmsiCheckboxField"]').click()
                time.sleep(2)
                self.driver.find_element(By.XPATH, '//*[@id="idSIButton9"]').click()
                time.sleep(2)
                logging.info("Logged in to Sharepoint\nPlease wait.....")

            except Exception as e:
                logging.error("Login failed: %s", e)
                return False

        def checkBuilder(bldr):
            # üßπ Always start by opening a fresh Kizuku page
            self.driver.execute_script("window.open('https://kizuku2.ctx.co.jp/logout', '_blank');")
            self.driver.switch_to.window(self.driver.window_handles[-1])
            time.sleep(3)

            WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="back"]'))).click()
            logging.info("baack to login page")
            # üìù Log builder name
            logging.info(f"Detected builder: {bldr}")

            # üß† Smartly decide which login to do
            if "„Ç±„Ç§„Ç¢„Ç§„Éó„É©„É≥„Éã„É≥„Ç∞" in bldr:
                try:
                    keiaiPlanningLogin()
                except Exception:
                    user = (
                        WebDriverWait(self.driver, 10)
                        .until(
                            EC.visibility_of_element_located(
                                (By.XPATH, '//*[@id="table-weekly-data"]/tbody/tr[1]/td[1]/div[1]')
                            )
                        )
                        .text
                    )
                    logging.info(f"Current logged user: {user}")
                    if "Èò™ÂíåNSK" in user:
                        WebDriverWait(self.driver, 10).until(
                            EC.element_to_be_clickable((By.XPATH, "//a[contains(string(),'Èò™ÂíåNSK')]"))
                        ).click()
                        WebDriverWait(self.driver, 10).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[@id="menu-navbar"]/ul[2]/li[4]/div/a[4]'))
                        ).click()
                        self.keiaiPlanningLogin()
                    else:
                        logging.info("Already Logged in Planning.")

            elif "ÈÅ†ÈâÑ„Éõ„Éº„É†" in bldr:
                try:
                    ÈÅ†ÈâÑ„Éõ„Éº„É†Login()
                except Exception:
                    user = (
                        WebDriverWait(self.driver, 10)
                        .until(
                            EC.visibility_of_element_located(
                                (By.XPATH, '//*[@id="table-weekly-data"]/tbody/tr[1]/td[1]/div[1]')
                            )
                        )
                        .text
                    )
                    logging.info(f"Current logged user: {user}")
                    if "„Ç®„Éå„Éª„Ç®„Çπ„Éª„Ç±„ÉºÂ∑•Ê•≠" in user:
                        WebDriverWait(self.driver, 10).until(
                            EC.element_to_be_clickable((By.XPATH, "//a[contains(string(),'„Ç®„Éå„Éª„Ç®„Çπ„Éª„Ç±„ÉºÂ∑•Ê•≠')]"))
                        ).click()
                        WebDriverWait(self.driver, 10).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[@id="menu-navbar"]/ul[2]/li[4]/div/a[5]'))
                        ).click()
                        ÈÅ†ÈâÑ„Éõ„Éº„É†Login()
                    else:
                        logging.info("Already Logged in ÈÅ†ÈâÑ„Éõ„Éº„É†.")
            else:
                try:
                    keiaistarFudousanLogin()
                except Exception:
                    user = (
                        WebDriverWait(self.driver, 10)
                        .until(
                            EC.visibility_of_element_located(
                                (By.XPATH, '//*[@id="table-weekly-data"]/tbody/tr[1]/td[1]/div[1]')
                            )
                        )
                        .text
                    )
                    logging.info(f"Current logged user: {user}")
                    if "Ëä≥Ë≥Ä„ÄÄÂíåÂâá" in user:
                        WebDriverWait(self.driver, 10).until(
                            EC.element_to_be_clickable((By.XPATH, "//a[contains(string(),'Ëä≥Ë≥Ä„ÄÄÂíåÂâá')]"))
                        ).click()
                        WebDriverWait(self.driver, 10).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[@id="menu-navbar"]/ul[2]/li[4]/div/a[4]'))
                        ).click()
                        keiaistarFudousanLogin()
                    else:
                        logging.info("Already Logged in Fudousan.")

        def fileUpload(upload_path, Ê°à‰ª∂Áï™Âè∑, Ê°à‰ª∂Âêç, „Éì„É´„ÉÄ„ÉºÂêç):
            username = "chirantindia1@nskkogyo.onmicrosoft.com"
            password = "Stupendocorona11"
            site_url = "https://nskkogyo.sharepoint.com/sites/kantou"

            ctx = ClientContext(site_url).with_credentials(UserCredential(username, password))

            builder_folder_map = {
                "ÈÅ†ÈâÑ„Éõ„Éº„É†Ôºà‰∏≠ÈÉ®„Éõ„Éº„É†„Çµ„Éº„Éì„ÇπÔºâ": "„ÅÇË°å/ÈÅ†ÈâÑ„Éõ„Éº„É†",
                "„Ç±„Ç§„Ç¢„Ç§„Çπ„Çø„Éº‰∏çÂãïÁî£": "„ÅãË°å/„Ç±„Ç§„Ç¢„Ç§„Çπ„Çø„Éº‰∏çÂãïÁî£",
                "„Ç±„Ç§„Ç¢„Ç§„Çπ„Çø„Éº‰∏çÂãïÁî£(Ê∫ñËÄêÁÅ´)": "„ÅãË°å/„Ç±„Ç§„Ç¢„Ç§„Çπ„Çø„Éº‰∏çÂãïÁî£",
                "TAKASUGIÔºà„Ç±„Ç§„Ç¢„Ç§„Çπ„Çø„ÉºÔºâ": "„ÅãË°å/„Ç±„Ç§„Ç¢„Ç§„Çπ„Çø„Éº‰∏çÂãïÁî£",
                "„Ç±„Ç§„Ç¢„Ç§„Éó„É©„É≥„Éã„É≥„Ç∞Ê†™Âºè‰ºöÁ§æ": "„ÅãË°å/„Ç±„Ç§„Ç¢„Ç§„Çπ„Çø„Éº‰∏çÂãïÁî£",
            }

            folder_base = builder_folder_map.get(„Éì„É´„ÉÄ„ÉºÂêç, "„ÅãË°å/„Ç±„Ç§„Ç¢„Ç§„Çπ„Çø„Éº‰∏çÂãïÁî£")

            combined_folder_name = f"{Ê°à‰ª∂Áï™Âè∑} {Ê°à‰ª∂Âêç}"
            alternate_combined_folder_name = f"{Ê°à‰ª∂Áï™Âè∑}\u3000{Ê°à‰ª∂Âêç}"  # full-width space

            sharepoint_folder_candidates = [
                PurePosixPath("/sites/kantou/Shared Documents", folder_base, combined_folder_name, "Ë≥áÊñô"),
                PurePosixPath("/sites/kantou/Shared Documents", folder_base, alternate_combined_folder_name, "Ë≥áÊñô"),
            ]

            Êñ∞Ë≥áÊñô_folder_candidates = [PurePosixPath(path, "Êñ∞Ë≥áÊñô") for path in sharepoint_folder_candidates]

            logging.info(f"üìÇ Upload path: {upload_path}")

            local_folder = Path(upload_path)
            if not any(local_folder.iterdir()):
                logging.warning(Fore.YELLOW + f"‚ö†Ô∏è Local upload folder is empty: {local_folder}")
                return

            # Step 1: Find Ë≥áÊñô folder
            Ë≥áÊñô_folder = try_open_folder(ctx, sharepoint_folder_candidates)
            if Ë≥áÊñô_folder is None:
                logging.error(Fore.RED + "‚ùå Ë≥áÊñô folder not found. Skipping upload.")
                return

            # Step 2: Find or create Êñ∞Ë≥áÊñô folder
            Êñ∞Ë≥áÊñô_folder = try_open_folder(ctx, Êñ∞Ë≥áÊñô_folder_candidates)
            if Êñ∞Ë≥áÊñô_folder is None:
                try:
                    Ë≥áÊñô_folder.folders.add("Êñ∞Ë≥áÊñô")
                    ctx.execute_query()
                    logging.info(Fore.GREEN + "‚úÖ Created Êñ∞Ë≥áÊñô folder inside Ë≥áÊñô.")
                    Êñ∞Ë≥áÊñô_folder = try_open_folder(ctx, Êñ∞Ë≥áÊñô_folder_candidates)
                except Exception as e:
                    logging.error(Fore.RED + f"‚ùå Could not create Êñ∞Ë≥áÊñô folder: {e}")
                    return

            # Step 3: Fetch existing files from Ë≥áÊñô folder
            try:
                ctx.load(Ë≥áÊñô_folder, ["Files"])
                ctx.execute_query()
                existing_files_set = set(normalize_name(f.properties["Name"]) for f in Ë≥áÊñô_folder.files)
                logging.info(Fore.CYAN + f"üîç Found {len(existing_files_set)} existing files in Ë≥áÊñô.")
            except Exception as e:
                logging.error(Fore.RED + f"‚ùå Failed to load files from Ë≥áÊñô folder: {e}")
                existing_files_set = set()

            files = list(local_folder.iterdir())
            logging.info(Fore.CYAN + f"üì¶ Preparing to upload {len(files)} files...")

            uploaded_count = 0
            skipped_count = 0
            failed_count = 0

            # Step 4: Upload with retries
            def upload_single_file(file_path):
                nonlocal uploaded_count, skipped_count, failed_count

                if file_path.is_file():
                    safe_file_name = sanitize_filename(file_path.name)
                    normalized_file_name = normalize_name(safe_file_name)

                    if normalized_file_name in existing_files_set:
                        logging.info(Fore.BLUE + f"‚è© Skipping duplicate file: {safe_file_name}")
                        skipped_count += 1
                        return

                    with open(file_path, "rb") as f:
                        file_content = f.read()

                    # Retry logic
                    max_retries = 5
                    for attempt in range(1, max_retries + 1):
                        try:
                            Êñ∞Ë≥áÊñô_folder.upload_file(safe_file_name, file_content)
                            ctx.execute_query()
                            logging.info(Fore.GREEN + f"‚úÖ Uploaded file: {safe_file_name}")
                            uploaded_count += 1
                            return
                        except Exception as e:
                            if attempt < max_retries:
                                logging.warning(Fore.YELLOW + f"‚ö†Ô∏è Retry {attempt} for {safe_file_name}: {e}")
                            else:
                                logging.error(Fore.RED + f"‚ùå Failed after {max_retries} retries: {safe_file_name}")
                                failed_count += 1
                                return

            with ThreadPoolExecutor(max_workers=5) as executor:
                futures = [executor.submit(upload_single_file, file_path) for file_path in files]
                for _ in tqdm(as_completed(futures), total=len(futures), desc="üöÄ Uploading files", dynamic_ncols=True):
                    pass

            # Step 5: Final Summary
            logging.info(Fore.CYAN + "\nüìÑ Upload Summary:")
            logging.info(Fore.GREEN + f"   ‚úÖ Uploaded: {uploaded_count} files")
            logging.info(Fore.BLUE + f"   ‚è© Skipped (duplicates): {skipped_count} files")
            logging.info(Fore.RED + f"   ‚ùå Failed: {failed_count} files")
            logging.info(Fore.MAGENTA + "üéâ Upload process completed successfully.\n")
            logging.info(Fore.GREEN + "Folder has been uploaded.")

        # --- Helper Functions ---
        def normalize_name(name):
            name = unicodedata.normalize("NFKC", name)
            name = name.replace("\u3000", " ")
            name = " ".join(name.split())
            return name.strip()

        def sanitize_filename(filename):
            invalid_chars = r'<>:"/\|?*'
            for char in invalid_chars:
                filename = filename.replace(char, "_")
            return filename

        def try_open_folder(ctx, folder_url_candidates):
            for url in folder_url_candidates:
                try:
                    folder = ctx.web.get_folder_by_server_relative_url(str(url))
                    ctx.load(folder)
                    ctx.execute_query()
                    logging.info(Fore.GREEN + f"‚úÖ Found folder: {url}")
                    return folder
                except Exception as e:
                    logging.warning(Fore.YELLOW + f"‚ö†Ô∏è Could not open {url}: {e}")
            logging.error(Fore.RED + "‚ùå All folder attempts failed.")
            return None

        def Access2(Ê°à‰ª∂Áï™Âè∑):
            # Switch to the correct window
            self.driver.switch_to.window(self.driver.window_handles[0])
            time.sleep(2)
            logging.info("Switched to access window")

            ÂèóÊ≥®‰∏ÄË¶ß_xpath = "/html/body/div[2]/div[1]/div/div/div/ul/li[4]/a"
            ÂèóÊ≥®‰∏ÄË¶ß = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, ÂèóÊ≥®‰∏ÄË¶ß_xpath)))
            ÂèóÊ≥®‰∏ÄË¶ß.click()
            logging.info("ÂèóÊ≥®‰∏ÄË¶ß clicked")
            time.sleep(2)

            Reset_xpath = "/html/body/div[2]/div[2]/div[2]/form/div/div/button[2]"
            Reset = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, Reset_xpath)))
            Reset.click()
            logging.info("Reset clicked")
            time.sleep(2)

            Anken_xpath = "/html/body/div[2]/div[2]/div[2]/form/div/table[3]/tbody/tr/td[1]/input"
            Anken = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, Anken_xpath)))
            Anken.send_keys(Ê°à‰ª∂Áï™Âè∑)
            logging.info("Anken number entered")
            time.sleep(2)

            Search_xpath = "/html/body/div[2]/div[2]/div[2]/form/div/div/button[1]"
            Search = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, Search_xpath)))
            Search.click()
            logging.info("Access Search clicked")
            time.sleep(2)

            try:
                ÂèÇÁÖß_xpath = "/html/body/div[2]/div[2]/div[2]/div[2]/div/div[3]/div[2]/div/table/tbody/tr/td[1]/input"
                ÂèÇÁÖß = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, ÂèÇÁÖß_xpath)))
                ÂèÇÁÖß.click()
                logging.info("ÂèÇÁÖß clicked")
            except NoSuchElementException:
                logging.error("ÂèÇÁÖß button not found. Moving to the next loop.")
                time.sleep(2)
                return False
            time.sleep(2)

            # Select the Âõ≥Èù¢
            Selectzumen = self.driver.find_element(By.NAME, "project[drawing]")
            Selectzumen.click()
            logging.info("Âõ≥Èù¢ clicked")
            time.sleep(1)

            # Select ÈÄÅ‰ªòÊ∏à
            Checkbox5 = self.driver.find_element(
                By.XPATH, "/html/body/div[2]/div[2]/div/div/form/div[2]/table[13]/tbody/tr[2]/td[6]/select/option[2]"
            )
            Checkbox5.click()
            logging.info("‰ΩúÂõ≥‰æùÈ†º clicked")
            time.sleep(2)

            Saveit_xpath = '(//*[@id="order_update"])[2]'
            Saveit = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, Saveit_xpath)))
            Saveit.click()
            logging.info("Save clicked")
            time.sleep(2)

            # Click on Ê°à‰ª∂‰∏ÄË¶ß
            ÂèóÊ≥®‰∏ÄË¶ß_xpath = "/html/body/div[2]/div[1]/div/div/div/ul/li[4]/a"
            ÂèóÊ≥®‰∏ÄË¶ß = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, ÂèóÊ≥®‰∏ÄË¶ß_xpath)))
            ÂèóÊ≥®‰∏ÄË¶ß.click()
            logging.info("ÂèóÊ≥®‰∏ÄË¶ß clicked")
            time.sleep(2)

        def Excelformating(Over):
            wb = load_workbook(Over)
            ws = wb.active
            column_widths = {"A": 15, "B": 40, "C": 10, "D": 15, "E": 100}
            for column, width in column_widths.items():
                ws.column_dimensions[column].width = width

            header_border = Border(
                left=Side(border_style="medium"),
                right=Side(border_style="medium"),
                top=Side(border_style="medium"),
                bottom=Side(border_style="medium"),
            )

            thin_border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )

            for cell in ws[1]:
                cell.border = header_border

            for row in ws.iter_rows(min_row=2, min_col=1, max_col=5):
                for cell in row:
                    cell.border = thin_border

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
                for cell in row:
                    if cell.value == "OK":
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    elif cell.value == "NG":
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            data_range = ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column)
            for row in data_range:
                for cell in row:
                    align = Alignment(horizontal="center", vertical="center")
                    cell.alignment = align

            # Adjust the page layout options
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0

            # Optional: Adjust margins if needed
            margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3)
            ws.page_margins = margins

            wb.save(Over)

        # Define the folder path where the ZIP files are located
        winrar_path = r"C:\Program Files\WinRAR\WinRAR.exe"  # Update this path if needed

        def zip_extract(abc_path):
            # List all ZIP files in the specified folder
            zip_files = [f for f in os.listdir(abc_path) if f.endswith(".zip")]

            if not zip_files:
                raise FileNotFoundError("No ZIP files found in the specified folder.")

            # Process each ZIP file
            for zip_filename in zip_files:
                zip_file_path = os.path.join(abc_path, zip_filename)

                # Define the extraction command
                extraction_command = [
                    winrar_path,  # Path to WinRAR executable
                    "x",  # Extract command
                    zip_file_path,  # The ZIP file to extract
                    abc_path,  # Destination folder
                ]

                # Run the command
                try:
                    subprocess.run(extraction_command, check=True)
                    print(f"Extracted contents of {zip_filename} to: {abc_path}")

                    # Remove the original ZIP file after extraction
                    os.remove(zip_file_path)
                    print(f"Removed ZIP file: {zip_filename}")

                except subprocess.CalledProcessError as e:
                    print(f"An error occurred while extracting {zip_filename}: {e}")

        def csv_to_excel(csv_files):
            for csv_file in csv_files:
                csv_file_path = os.path.join(rf"{os.getcwd()}", csv_file)
                try:
                    data = pd.read_csv(csv_file_path, encoding="CP932")
                    excel_file = os.path.join(rf"{os.getcwd()}", "Âõ≥Èù¢ÈÄÅ‰ªò„Éá„Éº„Çø.xlsx")
                    data.to_excel(excel_file, index=False, engine="openpyxl")
                    logging.info(f"File converted successfully and saved as {excel_file}")
                    os.remove(csv_file_path)
                except Exception as e:
                    logging.error(f"Failed to read and convert {csv_file}: {e}")
            return excel_file

        def clean_excel_data(excel_file):
            allowed_builders = [
                "„Ç±„Ç§„Ç¢„Ç§„Çπ„Çø„Éº‰∏çÂãïÁî£",
                "„Ç±„Ç§„Ç¢„Ç§„Çπ„Çø„Éº‰∏çÂãïÁî£(Ê∫ñËÄêÁÅ´)",
                "TAKASUGIÔºà„Ç±„Ç§„Ç¢„Ç§„Çπ„Çø„ÉºÔºâ",
                "„Ç±„Ç§„Ç¢„Ç§„Éó„É©„É≥„Éã„É≥„Ç∞Ê†™Âºè‰ºöÁ§æ",
            ]
            df = pd.read_excel(excel_file, sheet_name="Sheet1", dtype=str).astype(str).dropna()

            df_filtered = df[df["ÂæóÊÑèÂÖàÂêç"].isin(allowed_builders)]

            # Continue with any other processing with df_filtered
            logging.info(f"Filtered DataFrame has {len(df_filtered)} rows.")

            # Optionally, save the filtered DataFrame back to Excel if needed
            df_filtered.to_excel(excel_file, index=False, engine="openpyxl")

        def save_versioned_file(base_file):
            now = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_filename = f"Kistarkoushin_{now}.xlsx"
            new_filepath = os.path.join(os.getcwd(), new_filename)
            shutil.copy(base_file, new_filepath)
            logging.info(f"üìù Saved a versioned file: {new_filename}")

        def beautify_links(Over):
            wb = load_workbook(Over)
            ws = wb.active

            link_column = 5  # Column E (Ë≥áÊñô„É™„É≥„ÇØ)
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=link_column)
                if cell.value and cell.value.startswith("http"):
                    # Turn into hyperlink
                    cell.hyperlink = cell.value
                    cell.value = "üîó 365Link"
                    cell.font = Font(color="0000FF", underline="single")  # Blue underlined text

            wb.save(Over)
            logging.info("üîó Beautified all hyperlinks in Excel.")

        colorama_init(autoreset=True)

        empty_ankens_folder()
        time.sleep(1)

        clear_excel_data(Over)
        time.sleep(2)

        # Load Excel workbook
        wb = load_workbook(os.path.join(Over))
        sheet = wb.active

        handle_login()
        time.sleep(2)

        Accesslogin(Accessurl)
        time.sleep(2)

        Access(self.from_date, self.to_date)
        time.sleep(2)

        csv_files = [file for file in os.listdir(rf"{os.getcwd()}") if file.endswith(".csv")]
        excel_file = csv_to_excel(csv_files)

        clean_excel_data(excel_file)
        time.sleep(2)

        try:
            df = pd.read_excel(excel_file, sheet_name="Sheet1", dtype=str).astype(str).dropna()
            df["Á¢∫ÂÆöÁ¥çÊúü"] = df["Á¢∫ÂÆöÁ¥çÊúü"].astype(str).str.split(" ", expand=True)[0]
            logging.info(df)

            Ê°à‰ª∂Áï™Âè∑ = df["Ê°à‰ª∂Áï™Âè∑"]  # Project Number (Column C)
            Ê°à‰ª∂Âêç = df["Áâ©‰ª∂Âêç"]  # Project Name (Column ‚ÖÆ)
            Link = df["Ë≥áÊñô„É™„É≥„ÇØ"]
            „Éì„É´„ÉÄ„ÉºÂêç = df["ÂæóÊÑèÂÖàÂêç"]
            Á¥çÊúü = df["Á¢∫ÂÆöÁ¥çÊúü"]

            excellinenumber = 2

            for row_number in range(len(Ê°à‰ª∂Âêç)):
                logging.info(f"\n Current Run {Ê°à‰ª∂Áï™Âè∑[row_number]},{Ê°à‰ª∂Âêç[row_number]}\n")

                if pd.isna(Ê°à‰ª∂Âêç[row_number]) or pd.isna(„Éì„É´„ÉÄ„ÉºÂêç[row_number]):
                    break
                else:
                    sheet[f"A{excellinenumber}"].value = Ê°à‰ª∂Áï™Âè∑[row_number]
                    sheet[f"B{excellinenumber}"].value = Ê°à‰ª∂Âêç[row_number]
                    sheet[f"D{excellinenumber}"].value = Á¥çÊúü[row_number]
                    sheet[f"E{excellinenumber}"].value = Link[row_number]

                    abc_path = Path(os.path.join(os.getcwd(), folder_path, Ê°à‰ª∂Âêç[row_number], Docs1))

                    checkBuilder(„Éì„É´„ÉÄ„ÉºÂêç[row_number])
                    time.sleep(2)
                    Kizukuprocess(Ê°à‰ª∂Áï™Âè∑[row_number], Ê°à‰ª∂Âêç[row_number], excellinenumber, „Éì„É´„ÉÄ„ÉºÂêç[row_number])
                    time.sleep(2)
                    excellinenumber += 1
                    logging.info(f"current row number:{excellinenumber}")
                    wb.save(Over)
                    self.driver.switch_to.window(self.driver.window_handles[0])
                    time.sleep(2)
        finally:
            self.driver.quit()
            Excelformating(Over)
            beautify_links(Over)
            wb.save(Over)
            save_versioned_file(Over)
            logging.info("Task Completed")


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Kistarkoushin Bot")
        self.geometry("600x400")

        Label(self, text="Welcome to the Kistarkoushin Bot", font=("Roboto", -18, "bold")).pack(pady=10)

        Label(self, text="From Date:").place(relx=0.3, rely=0.4, anchor=CENTER)
        self.from_date_entry = DateEntry(self, date_pattern="yyyy/mm/dd")
        self.from_date_entry.place(relx=0.5, rely=0.4, anchor=CENTER)

        Label(self, text="To Date:").place(relx=0.3, rely=0.5, anchor=CENTER)
        self.to_date_entry = DateEntry(self, date_pattern="yyyy/mm/dd")
        self.to_date_entry.place(relx=0.5, rely=0.5, anchor=CENTER)

        self.start_button = Button(self, text="Start Bot", command=self.start_script, bg="#3290db", fg="white")
        self.start_button.place(relx=0.5, rely=0.6, anchor=CENTER)

        self.status_label = Label(self, text="", font=("Roboto", 12))
        self.status_label.place(relx=0.5, rely=0.7, anchor=CENTER)

    def start_script(self):
        self.start_button.config(state="disabled")
        self.status_label.config(text="üöÄ Bot is Running...", fg="green")
        self.from_date = self.from_date_entry.get_date()
        self.to_date = self.to_date_entry.get_date()
        threading.Thread(target=self.run_script).start()

    def run_script(self):
        try:
            bot = Kistarkoushin(self.from_date, self.to_date)
            bot.process_data()
            self.status_label.config(text="‚úÖ Bot Completed!", fg="blue")
        except Exception as e:
            logging.error(e)
            self.status_label.config(text=f"‚ùå Error: {e}", fg="red")
        finally:
            self.start_button.config(state="normal")


app = App()
app.mainloop()
